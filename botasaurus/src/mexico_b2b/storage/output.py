"""
Output export manager for JSON, CSV, Excel, and validation reports.
"""

import csv
import json
import os
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..config.settings import settings
from ..models.company import CanonicalCompany
from ..utils.logging import logger


class OutputManager:
    """Manages writing canonical datasets and validation reports to disk."""

    def __init__(self, output_dir: Optional[Path] = None):
        self.output_dir = output_dir or settings.OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write_json(
        self,
        companies: List[CanonicalCompany],
        filename: str = "mexico_companies.json",
        include_personal_contacts: Optional[bool] = None,
    ) -> Path:
        """Writes canonical company records to a formatted JSON array."""
        target_path = self.output_dir / filename
        include_contacts = (
            include_personal_contacts
            if include_personal_contacts is not None
            else settings.ENABLE_PERSONAL_CONTACT_FIELDS
        )

        data = [c.to_dict(include_personal_contacts=include_contacts) for c in companies]
        with open(target_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info("Generated JSON dataset", path=str(target_path), records=len(companies))
        return target_path

    def write_csv(
        self,
        companies: List[CanonicalCompany],
        filename: str = "mexico_companies.csv",
        include_personal_contacts: Optional[bool] = None,
    ) -> Path:
        """Writes flattened canonical company records to CSV."""
        target_path = self.output_dir / filename
        include_contacts = (
            include_personal_contacts
            if include_personal_contacts is not None
            else settings.ENABLE_PERSONAL_CONTACT_FIELDS
        )

        flat_rows = [c.to_flat_dict(include_personal_contacts=include_contacts) for c in companies]
        if not flat_rows:
            flat_rows = [{}]

        fieldnames = list(flat_rows[0].keys()) if flat_rows else []
        with open(target_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            if companies:
                writer.writerows(flat_rows)

        logger.info("Generated CSV dataset", path=str(target_path), records=len(companies))
        return target_path

    def write_xlsx(
        self,
        companies: List[CanonicalCompany],
        filename: str = "mexico_companies.xlsx",
        include_personal_contacts: Optional[bool] = None,
    ) -> Path:
        """Writes styled Excel workbook with summary and data sheets."""
        target_path = self.output_dir / filename
        include_contacts = (
            include_personal_contacts
            if include_personal_contacts is not None
            else settings.ENABLE_PERSONAL_CONTACT_FIELDS
        )

        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = "Companies"

        flat_rows = [c.to_flat_dict(include_personal_contacts=include_contacts) for c in companies]
        if not flat_rows:
            wb.save(target_path)
            return target_path

        headers = list(flat_rows[0].keys())
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border_thin = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        # Write header
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=header.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_num, row_data in enumerate(flat_rows, 2):
            for col_num, header in enumerate(headers, 1):
                val = row_data.get(header, "")
                cell = ws_data.cell(row=row_num, column=col_num, value=val)
                cell.border = border_thin
                if header in ("data_quality_score", "source_count"):
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col in ws_data.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

        wb.save(target_path)
        logger.info("Generated Excel dataset", path=str(target_path), records=len(companies))
        return target_path

    def write_people_json(
        self,
        people: List[Any],
        filename: str = "mexico_people.json",
        include_personal_contacts: Optional[bool] = None,
    ) -> Path:
        """Writes decision maker records to a formatted JSON array."""
        target_path = self.output_dir / filename
        include_contacts = (
            include_personal_contacts
            if include_personal_contacts is not None
            else settings.ENABLE_PERSONAL_CONTACT_FIELDS
        )

        data = [p.to_dict(include_personal_contacts=include_contacts) for p in people]
        with open(target_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info("Generated People JSON dataset", path=str(target_path), records=len(people))
        return target_path

    def write_people_csv(
        self,
        people: List[Any],
        filename: str = "mexico_people.csv",
        include_personal_contacts: Optional[bool] = None,
    ) -> Path:
        """Writes flattened decision-maker records to CSV."""
        target_path = self.output_dir / filename
        include_contacts = (
            include_personal_contacts
            if include_personal_contacts is not None
            else settings.ENABLE_PERSONAL_CONTACT_FIELDS
        )

        flat_rows = [p.to_flat_dict(include_personal_contacts=include_contacts) for p in people]
        if not flat_rows:
            flat_rows = [{}]

        fieldnames = list(flat_rows[0].keys()) if flat_rows else []
        with open(target_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            if people:
                writer.writerows(flat_rows)

        logger.info("Generated People CSV dataset", path=str(target_path), records=len(people))
        return target_path

    def write_people_xlsx(
        self,
        people: List[Any],
        filename: str = "mexico_people.xlsx",
        include_personal_contacts: Optional[bool] = None,
    ) -> Path:
        """Writes styled Excel workbook for decision makers."""
        target_path = self.output_dir / filename
        include_contacts = (
            include_personal_contacts
            if include_personal_contacts is not None
            else settings.ENABLE_PERSONAL_CONTACT_FIELDS
        )

        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = "Decision Makers"

        flat_rows = [p.to_flat_dict(include_personal_contacts=include_contacts) for p in people]
        if not flat_rows:
            wb.save(target_path)
            return target_path

        headers = list(flat_rows[0].keys())
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border_thin = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        # Write header
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=header.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_num, row_data in enumerate(flat_rows, 2):
            for col_num, header in enumerate(headers, 1):
                val = row_data.get(header, "")
                cell = ws_data.cell(row=row_num, column=col_num, value=val)
                cell.border = border_thin
                if header in ("email_status", "seniority_level", "email_confidence_score"):
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col in ws_data.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        wb.save(target_path)
        logger.info("Generated People Excel dataset", path=str(target_path), records=len(people))
        return target_path

    def write_validation_report(
        self,
        report_data: Dict[str, Any],
        filename: str = "validation_report.json",
    ) -> Path:
        """Writes detailed quality and validation metrics report."""
        target_path = self.output_dir / filename
        with open(target_path, "w", encoding="utf-8") as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
        logger.info("Generated validation report", path=str(target_path))
        return target_path

    def write_source_status(
        self,
        source_statuses: Dict[str, Any],
        filename: str = "source_status.json",
    ) -> Path:
        """Writes execution status and duration per source."""
        target_path = self.output_dir / filename
        with open(target_path, "w", encoding="utf-8") as f:
            json.dump(source_statuses, f, indent=2, ensure_ascii=False)
        return target_path


output_manager = OutputManager()
